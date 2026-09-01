from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

import open_table_connector.formulas as otf
import pytest
from open_table_connector.contract import (
    ConnectorError,
    ConnectorErrorCode,
    ProviderConfig,
    ProviderFactoryContext,
    ResolveContext,
    TableURI,
)
from open_table_connector.local_files import ExcelConnector, ExcelFormulaExtension
from open_table_connector.local_files.cli_adapter import (
    CsvCliAdapter,
    ExcelCliAdapter,
)
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

HASH = "sha256:" + "a" * 64


def _target(path: Path, sheet: str = "Model") -> otf.GridFormulaTarget:
    return otf.GridFormulaTarget(
        f"excel://{path.as_posix()}#sheet={sheet}",
        otf.WorksheetRef(name=sheet),
    )


def _workbook(path: Path) -> Path:
    workbook = Workbook()
    model = workbook.active
    model.title = "Model"
    model["A1"] = "value"
    model["B1"] = "=A1+1"
    model["C1"] = "=literal"
    model["C1"].data_type = "s"
    model["D1"] = 7
    model["A2"] = "keep"
    model["A2"].font = Font(bold=True, color="FF0000")
    model["A2"].comment = Comment("preserve this comment", "tester")
    validation = DataValidation(type="whole", operator="between", formula1="1", formula2="10")
    validation.add(model["A1"])
    model.add_data_validation(validation)
    model.sheet_properties.pageSetUpPr.fitToPage = True
    model.page_setup.fitToWidth = 1
    model.page_setup.fitToHeight = 0
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "hidden value"
    workbook.create_sheet("Last")
    workbook.defined_names.add(DefinedName("ModelValue", attr_text="'Model'!$A$1"))
    workbook.properties.title = "Formula preservation fixture"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    workbook.save(path)
    workbook.close()
    return path


def _bound(extension: ExcelFormulaExtension, path: Path) -> otf.BoundGridFormulaTarget:
    result = extension.bind_grid(otf.GridFormulaBindRequest(_target(path)))
    assert result.outcome is otf.FormulaOutcome.SUCCEEDED
    assert result.value is not None
    return result.value.target


def test_read_grid_uses_native_formula_cells_and_exact_worksheet_binding(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "model.xlsx")
    extension = ExcelFormulaExtension(ExcelConnector())

    binding_result = extension.bind_grid(otf.GridFormulaBindRequest(_target(path)))

    assert binding_result.outcome is otf.FormulaOutcome.SUCCEEDED
    assert binding_result.value is not None
    assert binding_result.value.target.worksheet.worksheet_id == "Model"
    assert binding_result.value.capabilities.details.dialects == (otf.EXCEL_A1,)
    assert binding_result.value.capabilities.details.max_cells_per_operation == 100_000
    assert binding_result.value.capabilities.details.max_expression_bytes == 8_192
    assert binding_result.value.capabilities.details.calculation_states == ()
    assert binding_result.value.capabilities.details.recalculation_scopes == ()

    result = extension.read_grid(otf.GridFormulaReadRequest(binding_result.value.target, "A1:D1"))

    assert result.outcome is otf.FormulaOutcome.SUCCEEDED
    assert result.value is not None
    assert [(cell.address, cell.expression.text) for cell in result.value.formulas] == [
        ("B1", "=A1+1")
    ]
    assert result.value.observed_revision.startswith("sha256:")
    assert result.receipts[0].calculation_state is None


@pytest.mark.parametrize(
    "uri_sheet, reference",
    [
        ("Other", otf.WorksheetRef(name="Model")),
        ("Model", otf.WorksheetRef(name="Missing")),
        ("Model", otf.WorksheetRef(worksheet_id="Missing")),
    ],
)
def test_bind_grid_rejects_conflicting_or_missing_worksheet_before_mutation(
    tmp_path: Path, uri_sheet: str, reference: otf.WorksheetRef
) -> None:
    path = _workbook(tmp_path / "model.xlsx")
    extension = ExcelFormulaExtension(ExcelConnector())

    result = extension.bind_grid(
        otf.GridFormulaBindRequest(
            otf.GridFormulaTarget(
                f"excel://{path.as_posix()}#sheet={uri_sheet}",
                reference,
            )
        )
    )

    assert result.outcome is otf.FormulaOutcome.REJECTED
    assert result.error is not None
    assert result.error.code is otf.FormulaErrorCode.TARGET_NOT_FOUND


def test_set_translates_top_left_formula_and_preserves_unrelated_workbook_objects(
    tmp_path: Path,
) -> None:
    path = _workbook(tmp_path / "model.xlsx")
    before_bytes = path.read_bytes()
    before_zip_names = set(ZipFile(path).namelist())
    extension = ExcelFormulaExtension(ExcelConnector())
    target = _bound(extension, path)
    before = extension.read_grid(otf.GridFormulaReadRequest(target, "A1:D1"))
    assert before.value is not None

    result = extension.set_grid(
        otf.GridFormulaSetRequest(
            target,
            "B2:D3",
            otf.FormulaExpression("=A2+$D$1", otf.EXCEL_A1),
            expected_revision=before.value.observed_revision,
            idempotency_key="set-1",
        )
    )

    assert result.outcome is otf.FormulaOutcome.SUCCEEDED
    assert result.commit is otf.FormulaCommitState.COMMITTED
    assert result.value is not None
    assert [
        (cell.address, cell.expression.text) for cell in result.value.formula_observation.formulas
    ] == [
        ("B2", "=A2+$D$1"),
        ("C2", "=B2+$D$1"),
        ("D2", "=C2+$D$1"),
        ("B3", "=A3+$D$1"),
        ("C3", "=B3+$D$1"),
        ("D3", "=C3+$D$1"),
    ]
    assert result.receipts[0].verification == "formula_text_readback"
    assert result.receipts[0].calculation_state is None
    assert result.receipts[0].value_observation_sha256 is None

    workbook = load_workbook(path, data_only=False)
    try:
        model = workbook["Model"]
        assert model["A2"].value == "keep"
        assert model["A2"].font.bold is True
        assert model["A2"].comment is not None
        assert model.data_validations.count == 1
        assert workbook.sheetnames == ["Model", "Hidden", "Last"]
        assert workbook["Hidden"].sheet_state == "hidden"
        assert workbook.defined_names["ModelValue"].attr_text == "'Model'!$A$1"
        assert workbook.calculation.fullCalcOnLoad is True
        assert workbook.calculation.forceFullCalc is True
        assert workbook.calculation.calcMode == "auto"
    finally:
        workbook.close()
    assert set(ZipFile(path).namelist()) == before_zip_names
    assert path.read_bytes() != before_bytes


def test_set_rejects_stale_revision_without_mutating_bytes(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "model.xlsx")
    extension = ExcelFormulaExtension(ExcelConnector())
    target = _bound(extension, path)
    before_bytes = path.read_bytes()

    result = extension.set_grid(
        otf.GridFormulaSetRequest(
            target,
            "B2",
            otf.FormulaExpression("=1", otf.EXCEL_A1),
            expected_revision=HASH,
        )
    )

    assert result.outcome is otf.FormulaOutcome.REJECTED
    assert result.error is not None
    assert result.error.code is otf.FormulaErrorCode.STALE_REVISION
    assert path.read_bytes() == before_bytes


def test_set_rejects_loss_of_unsupported_zip_parts_before_publication(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "custom.xlsx")
    with ZipFile(path, "a") as archive:
        archive.writestr("custom/unsupported.xml", b"<custom />")
    extension = ExcelFormulaExtension(ExcelConnector())
    target = _bound(extension, path)
    before = extension.read_grid(otf.GridFormulaReadRequest(target, "B1"))
    assert before.value is not None
    before_bytes = path.read_bytes()

    result = extension.set_grid(
        otf.GridFormulaSetRequest(
            target,
            "B2",
            otf.FormulaExpression("=1", otf.EXCEL_A1),
            expected_revision=before.value.observed_revision,
        )
    )

    assert result.outcome is otf.FormulaOutcome.FAILED
    assert result.error is not None
    assert result.error.code is otf.FormulaErrorCode.PROTOCOL_FAILURE
    assert path.read_bytes() == before_bytes


def test_limits_are_rejected_before_workbook_parse_and_after_formula_parse(tmp_path: Path) -> None:
    invalid = tmp_path / "invalid.xlsx"
    invalid.write_bytes(b"PK\x03\x04not a workbook")
    extension = ExcelFormulaExtension(ExcelConnector())
    target = otf.BoundGridFormulaTarget(
        f"excel://{invalid.as_posix()}#sheet=Model",
        otf.WorksheetRef(worksheet_id="Model"),
    )

    too_many = extension.read_grid(
        otf.GridFormulaReadRequest(target, "A1:B2", otf.FormulaResourceLimits(max_cells=3))
    )

    assert too_many.outcome is otf.FormulaOutcome.REJECTED
    assert too_many.error is not None
    assert too_many.error.code is otf.FormulaErrorCode.RESOURCE_LIMIT

    path = _workbook(tmp_path / "formula.xlsx")
    bound = _bound(extension, path)
    too_long = extension.read_grid(
        otf.GridFormulaReadRequest(bound, "B1", otf.FormulaResourceLimits(max_expression_bytes=3))
    )
    assert too_long.outcome is otf.FormulaOutcome.REJECTED
    assert too_long.error is not None
    assert too_long.error.code is otf.FormulaErrorCode.RESOURCE_LIMIT


def test_direct_excel_rejects_symlink_and_non_xlsx_payloads(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "model.xlsx")
    link = tmp_path / "link.xlsx"
    link.symlink_to(source)
    extension = ExcelFormulaExtension(ExcelConnector())

    result = extension.bind_grid(otf.GridFormulaBindRequest(_target(link)))
    assert result.outcome is otf.FormulaOutcome.REJECTED
    assert result.error is not None
    assert result.error.code is otf.FormulaErrorCode.INVALID_TARGET

    renamed = tmp_path / "renamed.xlsx"
    renamed.write_text("id,value\n1,2\n", encoding="utf-8")
    with pytest.raises(ConnectorError) as raised:
        ExcelConnector().resolve(TableURI(f"excel://{renamed}"), ResolveContext())
    assert raised.value.code is ConnectorErrorCode.INVALID_URI


def test_excel_has_no_calculated_value_or_recalculation_path_and_adapter_forwards_only_excel(
    tmp_path: Path,
) -> None:
    path = _workbook(tmp_path / "model.xlsx")
    extension = ExcelFormulaExtension(ExcelConnector())
    target = _bound(extension, path)

    values = extension.read_grid_values(otf.GridFormulaValueReadRequest(target, "A1"))
    recalculate = extension.recalculate_grid(
        otf.GridFormulaRecalculateRequest(target, otf.GridRecalculationScope.WORKBOOK)
    )

    assert values.outcome is otf.FormulaOutcome.REJECTED
    assert values.error is not None
    assert values.error.code is otf.FormulaErrorCode.UNSUPPORTED_CAPABILITY
    assert recalculate.outcome is otf.FormulaOutcome.REJECTED
    assert recalculate.error is not None
    assert recalculate.error.code is otf.FormulaErrorCode.UNSUPPORTED_CAPABILITY
    assert not hasattr(CsvCliAdapter, "formula_extension_for")
    adapter = ExcelCliAdapter(ExcelConnector(), ProviderFactoryContext(ProviderConfig("excel")))
    assert isinstance(adapter.formula_extension_for(), otf.CompositeFormulaConnectorExtension)


def test_same_idempotency_key_replays_without_second_publication(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "model.xlsx")
    extension = ExcelFormulaExtension(ExcelConnector())
    target = _bound(extension, path)
    before = extension.read_grid(otf.GridFormulaReadRequest(target, "A1"))
    assert before.value is not None
    request = otf.GridFormulaSetRequest(
        target,
        "B2",
        otf.FormulaExpression("=1", otf.EXCEL_A1),
        expected_revision=before.value.observed_revision,
        idempotency_key="same-key",
    )

    first = extension.set_grid(request)
    after_first = path.read_bytes()
    second = extension.set_grid(request)

    assert first.outcome is otf.FormulaOutcome.SUCCEEDED
    assert second == first
    assert path.read_bytes() == after_first

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
import xlwt

from open_table_connector.contract import ResolveContext, TableURI
from open_table_connector.contract.errors import ConnectorError, ConnectorErrorCode
from open_table_connector.local_files.reader import (
    LocalFilesConnector,
    LocalReadOptions,
    LocalTableReadRequest,
)
from open_table_connector.local_files.resolver import LocalFormat


@pytest.mark.parametrize(
    ("filename", "payload", "expected_format"),
    (
        ("orders.csv", "id\n1\n", LocalFormat.CSV),
        ("orders.tsv", "id\tamount\n1\t2\n", LocalFormat.CSV),
        ("orders.json", '[{"id": 1, "amount": 2.5}]', LocalFormat.JSON),
        ("orders.md", "| id |\n| --- |\n| 1 |\n", LocalFormat.MARKDOWN),
    ),
)
def test_local_files_facade_autodetects_supported_text_formats(
    tmp_path: Path, filename: str, payload: str, expected_format: LocalFormat
) -> None:
    source = tmp_path / filename
    source.write_text(payload, encoding="utf-8")
    connector = LocalFilesConnector()

    resolved = connector.resolve(TableURI(source.as_uri()), ResolveContext())
    result = connector.read_arrow(LocalTableReadRequest(TableURI(source.as_uri())))

    assert resolved.resource.format is expected_format
    assert result.table.num_rows == 1
    assert result.receipt.connector.connector_id == "local_files"


def test_local_files_facade_autodetects_tsv_separator(tmp_path: Path) -> None:
    source = tmp_path / "orders.tsv"
    source.write_text("id\tamount\n1\t2\n", encoding="utf-8")

    result = LocalFilesConnector().read_polars(
        LocalTableReadRequest(TableURI(source.as_uri()))
    )

    assert result.frame.to_dicts() == [{"id": "1", "amount": "2"}]


def test_local_files_facade_reads_json_array(tmp_path: Path) -> None:
    source = tmp_path / "orders.data"
    source.write_text('[{"id": 1, "amount": 2.5}]', encoding="utf-8")

    result = LocalFilesConnector().read_polars(
        LocalTableReadRequest(TableURI(source.as_uri()))
    )

    assert result.frame.to_dicts() == [{"id": 1, "amount": 2.5}]


def test_local_files_facade_reads_legacy_xls(tmp_path: Path) -> None:
    source = tmp_path / "orders.xls"
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("orders")
    worksheet.write(0, 0, "id")
    worksheet.write(0, 1, "amount")
    worksheet.write(0, 2, "source_date")
    worksheet.write(1, 0, "o-1")
    worksheet.write(1, 1, 2.5)
    worksheet.write(1, 2, "2026-06-01")
    worksheet.write(2, 0, "o-2")
    worksheet.write(2, 1, 3.5)
    worksheet.write(2, 2, 46175.0)
    workbook.save(str(source))

    result = LocalFilesConnector().read_polars(
        LocalTableReadRequest(TableURI(source.as_uri()))
    )

    assert result.frame.to_dicts() == [
        {"id": "o-1", "amount": 2.5, "source_date": "2026-06-01"},
        {"id": "o-2", "amount": 3.5, "source_date": "46175.0"},
    ]


def test_local_files_facade_reads_excel_and_preserves_compatibility_receipts(
    tmp_path: Path,
) -> None:
    source = tmp_path / "orders.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "orders"
    worksheet.append(["id"])
    worksheet.append(["1"])
    workbook.save(source)

    result = LocalFilesConnector().read_arrow(LocalTableReadRequest(TableURI(source.as_uri())))

    assert result.table.column_names == ["id"]
    assert result.receipt.connector.connector_id == "local_files"


@pytest.mark.parametrize(("filename", "payload"), (("orders.csv", "id\n1\n"), ("orders.md", "| id |\n| --- |\n| 1 |\n")))
def test_local_files_facade_rejects_sheet_option_for_non_excel_formats(
    tmp_path: Path, filename: str, payload: str
) -> None:
    source = tmp_path / filename
    source.write_text(payload, encoding="utf-8")

    with pytest.raises(ConnectorError) as raised:
        LocalFilesConnector().read_arrow(
            LocalTableReadRequest(
                TableURI(source.as_uri()),
                options=LocalReadOptions(sheet="orders"),
            )
        )

    assert raised.value.code is ConnectorErrorCode.INVALID_URI


def test_local_files_facade_inspection_honors_csv_read_options(tmp_path: Path) -> None:
    source = tmp_path / "orders.csv"
    source.write_bytes("name;amount\ncafé;1\n".encode("latin-1"))
    request = LocalTableReadRequest(
        TableURI(source.as_uri()),
        options=LocalReadOptions(separator=";", encoding="latin-1"),
    )

    inspection = LocalFilesConnector().inspect(request)

    assert inspection.columns == ("name", "amount")
    assert inspection.row_count == 1


def test_local_files_facade_inspection_honors_excel_read_options(tmp_path: Path) -> None:
    source = tmp_path / "orders.xlsx"
    workbook = Workbook()
    orders = workbook.active
    orders.title = "orders"
    orders.append(["ignored metadata"])
    orders.append(["id", "amount"])
    orders.append(["1", "2.50"])
    refunds = workbook.create_sheet("refunds")
    refunds.append(["ignored metadata"])
    refunds.append(["refund_id", "amount"])
    refunds.append(["r1", "1.00"])
    workbook.save(source)
    request = LocalTableReadRequest(
        TableURI(source.as_uri()),
        options=LocalReadOptions(sheet="refunds", header_row=2),
    )

    inspection = LocalFilesConnector().inspect(request)

    assert inspection.columns == ("refund_id", "amount")
    assert inspection.coordinate_convention.sheet == "refunds"
    assert inspection.coordinate_convention.header_rows == 2

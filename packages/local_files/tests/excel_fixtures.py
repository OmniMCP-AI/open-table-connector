from __future__ import annotations

import base64
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook

from packages.timeseries.tests.fixtures import ticks_table


def value_workbook(path: Path, *, sheet: str = "Ticks") -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    worksheet.append(ticks_table().column_names)
    table = ticks_table()
    temporal = {
        "ts": table["ts"].cast("int64").to_pylist(),
        "received_at": table["received_at"].cast("int64").to_pylist(),
    }
    columns = {
        name: table[name].to_pylist()
        for name in ("symbol", "venue", "price", "size")
    }
    for index in range(table.num_rows):
        worksheet.append(
            [
                _format_ns(temporal["ts"][index]),
                columns["symbol"][index],
                columns["venue"][index],
                columns["price"][index],
                columns["size"][index],
                _format_ns(temporal["received_at"][index]),
            ]
        )
    metadata = workbook.create_sheet("_otc_ts_schema")
    metadata.sheet_state = "hidden"
    metadata["A1"] = base64.b64encode(table.schema.serialize().to_pybytes()).decode("ascii")
    metadata["A2"] = sheet
    workbook.save(path)
    return path


def formula_workbook(path: Path, *, sheet: str = "Ticks") -> Path:
    value_workbook(path, sheet=sheet)
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    workbook[sheet]["D2"] = "=1+1"
    workbook.save(path)
    return path


def _format_ns(value: int) -> str:
    seconds, nanos = divmod(value, 1_000_000_000)
    return datetime.fromtimestamp(seconds, UTC).strftime("%Y-%m-%dT%H:%M:%S.") + f"{nanos:09d}Z"

"""Canonical Excel-to-Arrow materialization for the local-files Connector."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pyarrow as pa

from open_connectors.contract import ConnectorError, ConnectorErrorCode, ResourceLimits

from .csv_reader import _headers


def cell_text(value: Any) -> str | None:
    if value is None:
        return None
    return value if isinstance(value, str) else str(value)


def read_excel_arrow(
    path: Path,
    *,
    sheet: str | None,
    header_row: int,
    limits: ResourceLimits,
) -> tuple[pa.Table, str, tuple[str, ...]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ConnectorError(
            ConnectorErrorCode.UNSUPPORTED_CAPABILITY,
            "Excel reading requires openpyxl",
            {},
        ) from None

    try:
        workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
    except Exception as exc:
        raise ConnectorError(
            ConnectorErrorCode.EXECUTION_FAILED,
            "Excel workbook could not be opened",
            {"path": str(path), "reason": str(exc)},
        ) from None
    try:
        if sheet is None:
            worksheet = workbook.active
        else:
            try:
                worksheet = workbook[sheet]
            except KeyError:
                raise ConnectorError(
                    ConnectorErrorCode.INVALID_URI,
                    "Excel worksheet does not exist",
                    {"sheet": sheet, "worksheets": list(workbook.sheetnames)},
                ) from None
        rows = worksheet.iter_rows(min_row=header_row, values_only=True)
        try:
            header_values = [cell_text(value) or "" for value in next(rows)]
        except StopIteration:
            return pa.table({}), worksheet.title, tuple(workbook.sheetnames)
        names = _headers(header_values)
        data_rows: list[list[str | None]] = []
        for values in rows:
            row = list(values[: len(names)])
            if len(values) > len(names) and any(value is not None for value in values[len(names) :]):
                raise ConnectorError(
                    ConnectorErrorCode.EXECUTION_FAILED,
                    "Excel row has values beyond the header",
                    {"sheet": worksheet.title, "header_columns": len(names)},
                )
            normalized = [cell_text(value) for value in row]
            normalized.extend([None] * (len(names) - len(normalized)))
            if not any(value is not None for value in normalized):
                continue
            data_rows.append(normalized)
            if limits.max_rows is not None and len(data_rows) >= limits.max_rows:
                break
        columns = [
            pa.array([row[index] for row in data_rows], type=pa.large_string())
            for index in range(len(names))
        ]
        return pa.Table.from_arrays(columns, names=names), worksheet.title, tuple(workbook.sheetnames)
    except ConnectorError:
        raise
    except Exception as exc:
        raise ConnectorError(
            ConnectorErrorCode.EXECUTION_FAILED,
            "Excel worksheet could not be read",
            {"sheet": sheet, "reason": str(exc)},
        ) from None
    finally:
        workbook.close()

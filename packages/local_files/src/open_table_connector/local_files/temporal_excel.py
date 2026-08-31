"""Formula-safe portable temporal storage for governed Excel worksheets."""

from __future__ import annotations

import base64
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qsl, unquote, urlsplit

import pyarrow as pa

from open_table_connector.contract import (
    PROVIDER_EXCEL,
    SCHEME_MANAGED_XLSX,
    SCHEME_XLSX,
    TableURI,
)
from open_table_connector.timeseries import (
    ManagedAbortReceipt,
    ManagedAbortRequest,
    ManagedCommitReceipt,
    ManagedCommitRequest,
    ManagedReadbackRequest,
    ManagedReadbackResult,
    ManagedStageReceipt,
    ManagedStageRequest,
    PolarsTemporalExecutor,
    ResourceBounds,
    TemporalErrorCode,
    TemporalExecutionRequest,
    TemporalExecutionResult,
    TemporalExtensionError,
    TemporalTableDescriptor,
    TimestampPrecision,
)
from open_table_connector.timeseries.capabilities import (
    AGGREGATE_WINDOW,
    DESCRIBE,
    FILL,
    LOOKUP_ASOF,
    LOOKUP_LATEST,
    SCAN_RANGE,
    STORAGE_ABORT,
    STORAGE_COMMIT_IDEMPOTENT,
    STORAGE_READBACK_VERIFY,
    STORAGE_SNAPSHOT_READ,
    STORAGE_STAGE,
    STORAGE_VISIBILITY_ATOMIC,
)

from .managed_snapshots import ManagedSnapshotStore
from .identity import CONNECTOR_IDENTITY


_SCHEMA_SHEET = "_otc_ts_schema"


class ExcelManagedTemporalStore:
    def __init__(
        self,
        artifact_root: str | Path,
        descriptor: TemporalTableDescriptor,
        *,
        worksheet: str,
        clock=None,
        fault_injector=None,
    ) -> None:
        self.descriptor = _descriptor(descriptor)
        self.worksheet = _worksheet(worksheet)
        self.snapshots = ManagedSnapshotStore(
            artifact_root,
            descriptor,
            target_scheme=SCHEME_MANAGED_XLSX,
            extension=SCHEME_XLSX,
            encode_snapshot=lambda table: _encode_workbook(table, self.worksheet),
            decode_snapshot=lambda data: _decode_workbook(data, self.worksheet, descriptor),
            target_fragment=("sheet", self.worksheet),
            physical_target_validator=self._validate_physical_target,
            clock=clock,
            fault_injector=fault_injector,
        )

    @property
    def capabilities(self) -> tuple[str, ...]:
        return ExcelTemporalExecutor.CAPABILITIES

    def stage(self, request: ManagedStageRequest) -> ManagedStageReceipt:
        source = _direct_source(request.physical_target, self.worksheet, required=False)
        if source is not None:
            workbook_table = _decode_workbook(source.read_bytes(), self.worksheet, self.descriptor)
            _, artifact_table = self.snapshots._read_artifact(request)
            if not workbook_table.equals(artifact_table):
                raise TemporalExtensionError(
                    TemporalErrorCode.PROTOCOL_INVALID,
                    "governed Excel values do not match the staged Arrow artifact",
                    {},
                )
        return self.snapshots.stage_artifact(request)

    def commit(self, request: ManagedCommitRequest) -> ManagedCommitReceipt:
        return self.snapshots.publish_snapshot(request)

    def readback(self, request: ManagedReadbackRequest) -> ManagedReadbackResult:
        return self.snapshots.readback_snapshot(request)

    def abort(self, request: ManagedAbortRequest) -> ManagedAbortReceipt:
        return self.snapshots.abort_stage(request)

    def read_snapshot(
        self,
        target: TableURI,
        snapshot_reference: str,
        bounds: ResourceBounds,
    ) -> pa.Table:
        return self.snapshots.read_snapshot(target, snapshot_reference, bounds)

    def resolve_snapshot(self, target: TableURI, snapshot_reference: str) -> Path:
        return self.snapshots.resolve_snapshot(target, snapshot_reference)

    def recover(self, target: TableURI) -> None:
        self.snapshots.recover(target)

    def _validate_physical_target(self, target: TableURI) -> None:
        if target.scheme == "managed+xlsx":
            _managed_namespace(target, self.worksheet)
            return
        source = _direct_source(target, self.worksheet, required=True)
        _decode_workbook(source.read_bytes(), self.worksheet, self.descriptor)


class _ExcelTemporalSource:
    def __init__(self, table: pa.Table, descriptor: TemporalTableDescriptor) -> None:
        self._table = table
        self.descriptor = descriptor

    def read_bounded(self, target, projection, predicates, bounds) -> pa.Table:
        del target, predicates, bounds
        return self._table.select(projection)


class ExcelTemporalExecutor:
    CAPABILITIES = (
        DESCRIBE,
        SCAN_RANGE,
        LOOKUP_LATEST,
        LOOKUP_ASOF,
        AGGREGATE_WINDOW,
        FILL,
        STORAGE_STAGE,
        STORAGE_COMMIT_IDEMPOTENT,
        STORAGE_SNAPSHOT_READ,
        STORAGE_READBACK_VERIFY,
        STORAGE_VISIBILITY_ATOMIC,
        STORAGE_ABORT,
    )

    def __init__(
        self,
        descriptor: TemporalTableDescriptor,
        *,
        worksheet: str,
        managed_store: ExcelManagedTemporalStore | None = None,
    ) -> None:
        self.descriptor = _descriptor(descriptor)
        self.worksheet = _worksheet(worksheet)
        if managed_store is not None and managed_store.worksheet != self.worksheet:
            raise ValueError("executor and managed store worksheets must match")
        self.managed_store = managed_store

    def execute(self, request: TemporalExecutionRequest) -> TemporalExecutionResult:
        if not isinstance(request, TemporalExecutionRequest):
            raise TypeError("request must be a TemporalExecutionRequest")
        if request.target.scheme == "managed+xlsx":
            _managed_namespace(request.target, self.worksheet)
            if self.managed_store is None or request.snapshot_reference is None:
                raise TemporalExtensionError(
                    TemporalErrorCode.SNAPSHOT_UNAVAILABLE,
                    "managed Excel execution requires an addressed snapshot",
                    {},
                )
            table = self.managed_store.read_snapshot(
                request.target,
                request.snapshot_reference,
                request.plan.resource_bounds,
            )
        else:
            path = _direct_source(request.target, self.worksheet, required=True)
            size = path.stat().st_size
            if size > request.plan.resource_bounds.max_bytes:
                raise TemporalExtensionError(
                    TemporalErrorCode.RESOURCE_LIMIT_EXCEEDED,
                    "Excel source exceeds max_bytes",
                    {"bytes": size},
                )
            table = _decode_workbook(path.read_bytes(), self.worksheet, self.descriptor)
            if table.num_rows > request.plan.resource_bounds.max_rows:
                raise TemporalExtensionError(
                    TemporalErrorCode.RESOURCE_LIMIT_EXCEEDED,
                    "Excel source exceeds max_rows",
                    {"rows": table.num_rows},
                )
        return PolarsTemporalExecutor(
            _ExcelTemporalSource(table, self.descriptor), connector_identity=CONNECTOR_IDENTITY
        ).execute(request)


def _encode_workbook(table: pa.Table, worksheet_name: str) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_name
    _append_row(worksheet, list(table.column_names))
    columns = []
    for field, column in zip(table.schema, table.columns, strict=True):
        if pa.types.is_timestamp(field.type):
            raw = column.cast(pa.int64()).to_pylist()
            scale = {"s": 1_000_000_000, "ms": 1_000_000, "us": 1_000, "ns": 1}[
                field.type.unit
            ]
            columns.append(
                [None if value is None else _format_ns(value * scale) for value in raw]
            )
        elif pa.types.is_list(field.type) or pa.types.is_struct(field.type) or pa.types.is_map(field.type):
            raise TemporalExtensionError(
                TemporalErrorCode.PROTOCOL_INVALID,
                "nested Arrow values are unsupported in governed Excel snapshots",
                {"field": field.name},
            )
        else:
            columns.append(column.to_pylist())
    for row_index in range(table.num_rows):
        _append_row(
            worksheet,
            [columns[column_index][row_index] for column_index in range(table.num_columns)],
        )
    metadata = workbook.create_sheet(_SCHEMA_SHEET)
    metadata.sheet_state = "hidden"
    metadata["A1"] = base64.b64encode(table.schema.serialize().to_pybytes()).decode("ascii")
    metadata["A2"] = worksheet_name
    destination = BytesIO()
    workbook.save(destination)
    workbook.close()
    return destination.getvalue()


def _append_row(worksheet, values) -> None:
    worksheet.append(values)
    for cell in worksheet[worksheet.max_row]:
        if isinstance(cell.value, str):
            cell.data_type = "s"


def _decode_workbook(
    data: bytes,
    worksheet_name: str,
    descriptor: TemporalTableDescriptor,
) -> pa.Table:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except Exception as exc:
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel workbook could not be opened",
            {},
        ) from exc
    try:
        if worksheet_name not in workbook.sheetnames:
            raise TemporalExtensionError(
                TemporalErrorCode.PROTOCOL_INVALID,
                "governed Excel worksheet does not exist",
                {"worksheet": worksheet_name},
            )
        worksheet = workbook[worksheet_name]
        for row in worksheet.iter_rows():
            if any(cell.data_type == "f" for cell in row):
                raise TemporalExtensionError(
                    TemporalErrorCode.PROTOCOL_INVALID,
                    "formula is forbidden in the governed Excel worksheet",
                    {"worksheet": worksheet_name},
                )
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = tuple("" if value is None else str(value) for value in next(rows))
        except StopIteration:
            raise TemporalExtensionError(
                TemporalErrorCode.PROTOCOL_INVALID,
                "governed Excel worksheet is empty",
                {},
            ) from None
        if not header or any(not name for name in header) or len(set(header)) != len(header):
            raise TemporalExtensionError(
                TemporalErrorCode.PROTOCOL_INVALID,
                "governed Excel worksheet requires unique non-empty headers",
                {},
            )
        values = []
        for source in rows:
            row = list(source[: len(header)])
            if len(source) > len(header) and any(value is not None for value in source[len(header) :]):
                raise TemporalExtensionError(
                    TemporalErrorCode.PROTOCOL_INVALID,
                    "governed Excel row extends beyond its header",
                    {},
                )
            row.extend([None] * (len(header) - len(row)))
            if any(value is not None for value in row):
                values.append(row)
        schema = _embedded_schema(workbook, worksheet_name)
        table = _table_from_rows(header, values, schema, descriptor)
        _validate_descriptor_columns(table, descriptor)
        return table
    finally:
        workbook.close()


def _embedded_schema(workbook, worksheet_name: str) -> pa.Schema | None:
    if _SCHEMA_SHEET not in workbook.sheetnames:
        return None
    metadata = workbook[_SCHEMA_SHEET]
    if metadata["A2"].value != worksheet_name or not isinstance(metadata["A1"].value, str):
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel snapshot schema metadata does not match the governed worksheet",
            {},
        )
    try:
        return pa.ipc.read_schema(pa.BufferReader(base64.b64decode(metadata["A1"].value, validate=True)))
    except Exception as exc:
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel snapshot schema metadata is invalid",
            {},
        ) from exc


def _table_from_rows(header, rows, schema, descriptor) -> pa.Table:
    if schema is not None and tuple(schema.names) != header:
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel snapshot headers do not match schema metadata",
            {},
        )
    temporal = {descriptor.time_field, descriptor.ingestion_time_field}
    unit = {
        TimestampPrecision.SECOND: "s",
        TimestampPrecision.MILLISECOND: "ms",
        TimestampPrecision.MICROSECOND: "us",
        TimestampPrecision.NANOSECOND: "ns",
    }[descriptor.precision]
    arrays = []
    for index, name in enumerate(header):
        column = [row[index] for row in rows]
        expected_type = schema.field(name).type if schema is not None else None
        if name in temporal:
            timestamp_type = expected_type or pa.timestamp(unit, tz=descriptor.timezone)
            inferred = pa.array(column)
            arrays.append(inferred.cast(timestamp_type))
        elif expected_type is not None:
            arrays.append(pa.array(column, type=expected_type))
        else:
            inferred = pa.array(column)
            if pa.types.is_string(inferred.type):
                inferred = pa.array(column, type=pa.large_string())
            arrays.append(inferred)
    return pa.Table.from_arrays(arrays, names=header)


def _validate_descriptor_columns(table: pa.Table, descriptor: TemporalTableDescriptor) -> None:
    required = {
        descriptor.time_field,
        *descriptor.series_key_fields,
        *descriptor.tag_fields,
        *descriptor.value_fields,
    }
    if descriptor.ingestion_time_field is not None:
        required.add(descriptor.ingestion_time_field)
    missing = sorted(required - set(table.column_names))
    if missing:
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel worksheet does not satisfy its temporal descriptor",
            {"missing_fields": missing},
        )


def _target_parts(target: TableURI, worksheet: str):
    parsed = urlsplit(target.value)
    try:
        fragments = parse_qsl(parsed.fragment, keep_blank_values=True, strict_parsing=True)
    except ValueError as exc:
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel target fragment is invalid",
            {},
        ) from exc
    path = Path(unquote(parsed.path))
    if (
        parsed.netloc not in {"", "localhost"}
        or parsed.query
        or fragments != [("sheet", worksheet)]
        or not path.is_absolute()
        or ".." in path.parts
        or not path.name
    ):
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "Excel target must bind one absolute path and governed worksheet",
            {},
        )
    return path


def _direct_source(target: TableURI, worksheet: str, *, required: bool) -> Path | None:
    if target.scheme not in {SCHEME_XLSX, PROVIDER_EXCEL}:
        if required:
            raise TemporalExtensionError(
                TemporalErrorCode.PROTOCOL_INVALID,
                "direct Excel target requires xlsx or excel scheme",
                {"scheme": target.scheme},
            )
        return None
    path = _target_parts(target, worksheet)
    if not path.is_file() or path.is_symlink():
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "direct Excel target must be a regular non-symlink file",
            {},
        )
    return path


def _managed_namespace(target: TableURI, worksheet: str) -> Path:
    if target.scheme != "managed+xlsx":
        raise TemporalExtensionError(
            TemporalErrorCode.PROTOCOL_INVALID,
            "managed Excel target requires managed+xlsx scheme",
            {"scheme": target.scheme},
        )
    return _target_parts(target, worksheet)


def _descriptor(value: TemporalTableDescriptor) -> TemporalTableDescriptor:
    if not isinstance(value, TemporalTableDescriptor):
        raise TypeError("descriptor must be a TemporalTableDescriptor")
    return value


def _worksheet(value: str) -> str:
    if not isinstance(value, str) or not value.strip() or len(value) > 31:
        raise ValueError("worksheet must be a non-empty Excel worksheet name")
    if value == _SCHEMA_SHEET or any(character in value for character in "[]:*?/\\"):
        raise ValueError("worksheet is reserved or contains invalid Excel characters")
    return value


def _format_ns(value: int) -> str:
    seconds, nanos = divmod(value, 1_000_000_000)
    return datetime.fromtimestamp(seconds, UTC).strftime("%Y-%m-%dT%H:%M:%S.") + f"{nanos:09d}Z"


__all__ = ["ExcelManagedTemporalStore", "ExcelTemporalExecutor"]

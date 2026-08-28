from openpyxl import load_workbook
import pyarrow as pa
import pytest

from open_table_connector.contract import ConnectorError, ConnectorErrorCode, ResourceLimits
from open_table_connector.local_files.excel_reader import read_excel_arrow
from open_table_connector.local_files.excel_writer import write_excel


def test_excel_writer_writes_headers_and_rows_to_named_sheet(tmp_path) -> None:
    destination = tmp_path / "orders.xlsx"
    table = pa.table({"id": ["1", "2"], "amount": ["10", None]})

    write_excel(table, destination, sheet="Orders")

    workbook = load_workbook(destination, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["Orders"]
        assert list(workbook["Orders"].values) == [
            ("id", "amount"),
            ("1", "10"),
            ("2", None),
        ]
    finally:
        workbook.close()


def test_excel_writer_uses_default_sheet_name_and_writes_data(tmp_path) -> None:
    destination = tmp_path / "orders.xlsx"
    table = pa.table({"id": ["1"], "amount": ["10"]})

    write_excel(table, destination)

    workbook = load_workbook(destination, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["Sheet1"]
        assert list(workbook["Sheet1"].values) == [
            ("id", "amount"),
            ("1", "10"),
        ]
    finally:
        workbook.close()


def test_excel_writer_round_trips_formula_prefixed_headers_and_values_as_text(
    tmp_path,
) -> None:
    destination = tmp_path / "formulas.xlsx"
    table = pa.table({"=header": ["=1+1"]})

    write_excel(table, destination)

    workbook = load_workbook(destination, read_only=True, data_only=False)
    try:
        worksheet = workbook.active
        assert worksheet["A1"].value == "=header"
        assert worksheet["A1"].data_type == "s"
        assert worksheet["A2"].value == "=1+1"
        assert worksheet["A2"].data_type == "s"
    finally:
        workbook.close()

    round_trip, _, _ = read_excel_arrow(
        destination,
        sheet=None,
        header_row=1,
        limits=ResourceLimits(),
    )
    assert round_trip.to_pylist() == [{"=header": "=1+1"}]


def test_excel_writer_preserves_numeric_boolean_and_null_cells(tmp_path) -> None:
    destination = tmp_path / "typed.xlsx"
    table = pa.table(
        {
            "count": pa.array([7], type=pa.int64()),
            "active": pa.array([True], type=pa.bool_()),
            "missing": pa.array([None], type=pa.null()),
        }
    )

    write_excel(table, destination)

    workbook = load_workbook(destination, read_only=True, data_only=True)
    try:
        assert list(workbook.active.values) == [
            ("count", "active", "missing"),
            (7, True, None),
        ]
    finally:
        workbook.close()


def test_excel_writer_maps_file_errors_to_connector_error(tmp_path) -> None:
    destination = tmp_path / "missing" / "orders.xlsx"

    with pytest.raises(ConnectorError) as error:
        write_excel(pa.table({"id": ["1"]}), destination)

    assert error.value.code is ConnectorErrorCode.EXECUTION_FAILED
    assert error.value.safe_details["path"] == str(destination)

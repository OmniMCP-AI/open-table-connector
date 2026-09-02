"""Direct, value-safe Formula support for existing Excel workbooks."""

from __future__ import annotations

import hashlib
import os
import posixpath
import stat
import tempfile
import threading
from collections import OrderedDict
from contextlib import contextmanager, suppress
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import open_table_connector.formulas as otf
from open_table_connector.contract import PROVIDER_EXCEL, ConnectorError, ResolveContext

from .excel_connector import ExcelConnector

_MAX_CELLS = 100_000
_MAX_EXPRESSION_BYTES = 8_192
_DEFAULT_LEDGER_LIMIT = 1_024
_COMPLETED_CACHE_LIMIT = _DEFAULT_LEDGER_LIMIT
_UNSUPPORTED_RECALC_MESSAGE = "direct Excel does not expose calculated values or recalculation"


class _LimitFailure(ValueError):
    def __init__(self, message: str, limit: int) -> None:
        super().__init__(message)
        self.limit = limit


class _ProtocolFailure(ValueError):
    pass


class _PublicationFailure(OSError):
    def __init__(self, message: str, *, replaced: bool) -> None:
        super().__init__(message)
        self.replaced = replaced


def _hash_bytes(data: bytes) -> str:
    return "sha256:" + hashlib.sha256(data).hexdigest()


def _success(value: object, receipts: tuple[object, ...] = ()) -> otf.FormulaExtensionResult[Any]:
    return otf.FormulaExtensionResult(
        value=value,
        outcome=otf.FormulaOutcome.SUCCEEDED,
        commit=otf.FormulaCommitState.NOT_APPLICABLE,
        verification=otf.FormulaVerificationState.PASSED,
        receipts=receipts,
    )


def _rejected(
    code: otf.FormulaErrorCode,
    message: str,
    details: dict[str, object] | None = None,
) -> otf.FormulaExtensionResult[Any]:
    return otf.FormulaExtensionResult(
        value=None,
        outcome=otf.FormulaOutcome.REJECTED,
        commit=otf.FormulaCommitState.NOT_STARTED,
        verification=otf.FormulaVerificationState.SKIPPED,
        receipts=(),
        error=otf.FormulaExtensionErrorInfo(code, message, details or {}),
    )


def _failed(
    code: otf.FormulaErrorCode,
    message: str,
    *,
    commit: otf.FormulaCommitState = otf.FormulaCommitState.NOT_COMMITTED,
    verification: otf.FormulaVerificationState = otf.FormulaVerificationState.SKIPPED,
    receipts: tuple[object, ...] = (),
) -> otf.FormulaExtensionResult[Any]:
    return otf.FormulaExtensionResult(
        value=None,
        outcome=otf.FormulaOutcome.FAILED,
        commit=commit,
        verification=verification,
        receipts=receipts,
        error=otf.FormulaExtensionErrorInfo(code, message, {}),
    )


def _unknown(message: str) -> otf.FormulaExtensionResult[Any]:
    return otf.FormulaExtensionResult(
        value=None,
        outcome=otf.FormulaOutcome.UNKNOWN,
        commit=otf.FormulaCommitState.UNKNOWN,
        verification=otf.FormulaVerificationState.UNAVAILABLE,
        receipts=(),
        error=otf.FormulaExtensionErrorInfo(otf.FormulaErrorCode.UNCERTAIN_MUTATION, message, {}),
    )


class ExcelFormulaExtension(otf.GridFormulaConnectorExtension):
    """Formula adapter for direct ``excel://`` targets only."""

    def __init__(self, connector: ExcelConnector | None = None) -> None:
        self._connector = connector or ExcelConnector()
        self._bindings: dict[tuple[str, str], str] = {}
        self._ledger = otf.FormulaIdempotencyLedger(limit=_DEFAULT_LEDGER_LIMIT)
        self._completed_limit = _COMPLETED_CACHE_LIMIT
        self._completed: OrderedDict[
            str, otf.FormulaExtensionResult[otf.FormulaMutation]
        ] = OrderedDict()
        self._lock = threading.RLock()

    def bind_grid(
        self, request: otf.GridFormulaBindRequest
    ) -> otf.FormulaExtensionResult[otf.GridFormulaBinding]:
        try:
            path, uri_sheet = self._resolve_path(request.target.grid)
            self._assert_target_file(path)
            data = self._read_bytes(path)
            self._validate_zip(data)
            from openpyxl import load_workbook

            workbook = load_workbook(
                BytesIO(data), data_only=False, read_only=False, keep_links=True
            )
            try:
                worksheet = self._match_worksheet(
                    workbook.sheetnames, request.target.worksheet, uri_sheet
                )
                worksheet_id = worksheet
            finally:
                workbook.close()
            details = self._details()
            target = otf.BoundGridFormulaTarget(
                request.target.grid.value,
                otf.WorksheetRef(worksheet_id=worksheet_id),
            )
            with self._lock:
                self._bindings[(target.grid.value, worksheet_id)] = worksheet
            binding = otf.GridFormulaBinding(
                target=target,
                capabilities=otf.FormulaCapabilitySet((otf.GRID_READ, otf.GRID_SET), details),
                observed_revision=_hash_bytes(data),
            )
            return _success(binding)
        except _TargetFailure as exc:
            return _rejected(exc.code, str(exc))
        except _LimitFailure as exc:
            return _rejected(otf.FormulaErrorCode.RESOURCE_LIMIT, str(exc), {"limit": exc.limit})
        except (BadZipFile, OSError, _ProtocolFailure, TypeError, ValueError, KeyError):
            return _failed(
                otf.FormulaErrorCode.PROTOCOL_FAILURE, "Excel workbook could not be opened"
            )
        except Exception:
            return _failed(otf.FormulaErrorCode.EXECUTION_FAILED, "Excel worksheet binding failed")

    def read_grid(
        self, request: otf.GridFormulaReadRequest
    ) -> otf.FormulaExtensionResult[otf.GridFormulaObservation]:
        try:
            path, worksheet_name = self._bound_path(request.target)
            rectangle = self._validated_range(request.cell_range, request.limits)
            data = self._read_bytes(path)
            self._check_response_limit(data, request.limits)
            observation = self._read_observation(data, worksheet_name, rectangle, request.limits)
            receipt = otf.FormulaReceiptDetails.for_grid_read(
                target=request.target.grid.value,
                selector=request.cell_range,
                capability=otf.GRID_READ.to_reference(),
                dialect=otf.EXCEL_A1,
                observation_sha256=otf.formula_observation_hash(observation),
                observed_count=len(observation.formulas),
                revision_after=observation.observed_revision,
            )
            return _success(observation, (receipt,))
        except _TargetFailure as exc:
            return _rejected(exc.code, str(exc))
        except _LimitFailure as exc:
            return _rejected(otf.FormulaErrorCode.RESOURCE_LIMIT, str(exc), {"limit": exc.limit})
        except (BadZipFile, OSError, _ProtocolFailure, TypeError, ValueError, KeyError):
            return _failed(
                otf.FormulaErrorCode.PROTOCOL_FAILURE,
                "Excel formula read returned an invalid workbook",
            )
        except Exception:
            return _failed(otf.FormulaErrorCode.EXECUTION_FAILED, "Excel formula read failed")

    def read_grid_values(
        self, request: otf.GridFormulaValueReadRequest
    ) -> otf.FormulaExtensionResult[otf.GridFormulaValueObservation]:
        del request
        return _rejected(otf.FormulaErrorCode.UNSUPPORTED_CAPABILITY, _UNSUPPORTED_RECALC_MESSAGE)

    def recalculate_grid(
        self, request: otf.GridFormulaRecalculateRequest
    ) -> otf.FormulaExtensionResult[otf.RecalculationObservation]:
        del request
        return _rejected(otf.FormulaErrorCode.UNSUPPORTED_CAPABILITY, _UNSUPPORTED_RECALC_MESSAGE)

    def set_grid(
        self, request: otf.GridFormulaSetRequest
    ) -> otf.FormulaExtensionResult[otf.FormulaMutation]:
        context: tuple[str, str, str] | None = None
        replaced = False
        try:
            path, worksheet_name = self._bound_path(request.target)
            rectangle = self._validated_range(request.cell_range, request.limits)
            if request.expression.dialect != otf.EXCEL_A1:
                return _rejected(
                    otf.FormulaErrorCode.INVALID_FORMULA, "Excel formula dialect is invalid"
                )
            if not request.expression.text.startswith("="):
                return _rejected(
                    otf.FormulaErrorCode.INVALID_FORMULA, "Excel formulas must start with '='"
                )
            expression_limit = self._expression_limit(request.limits)
            if request.expression.byte_count > expression_limit:
                raise _LimitFailure(
                    "formula expression exceeds the configured byte limit", expression_limit
                )
            target_identity = f"{path.resolve()}\0{worksheet_name}"
            target_hash = _hash_bytes(target_identity.encode("utf-8"))
            selector_hash = _hash_bytes(otf.GRID_SET.to_reference().encode("utf-8"))
            payload_hash = _hash_bytes(
                f"{self._range_text(rectangle)}\0{request.expression.dialect}\0{request.expression.sha256}\0{request.expected_revision}".encode()
            )
            context = (target_hash, selector_hash, payload_hash)
            lock_path = Path(f"{path}.otc.lock")
            with self._target_lock(path, lock_path):
                current = self._read_bytes(path)
                self._check_response_limit(current, request.limits)
                before_revision = _hash_bytes(current)
                if request.idempotency_key is not None:
                    with self._lock:
                        decision = self._ledger.begin(
                            connector_id=PROVIDER_EXCEL,
                            capability=otf.GRID_SET.to_reference(),
                            target_hash=target_hash,
                            selector_hash=selector_hash,
                            idempotency_key=request.idempotency_key,
                            payload_hash=payload_hash,
                        )
                    if decision.disposition is otf.FormulaIdempotencyDisposition.CONFLICT:
                        return _rejected(
                            otf.FormulaErrorCode.IDEMPOTENCY_CONFLICT,
                            "formula idempotency key conflicts with a prior request",
                        )
                    if decision.disposition is otf.FormulaIdempotencyDisposition.IN_FLIGHT:
                        return _rejected(
                            otf.FormulaErrorCode.IDEMPOTENCY_CONFLICT,
                            "formula idempotency key is already in flight",
                        )
                    if decision.disposition is otf.FormulaIdempotencyDisposition.UNKNOWN:
                        return _unknown("formula mutation remains uncertain")
                    if decision.disposition is otf.FormulaIdempotencyDisposition.REPLAY:
                        with self._lock:
                            cached = self._completed.get(decision.operation_hash or "")
                            if cached is not None:
                                self._completed.move_to_end(decision.operation_hash or "")
                        return (
                            cached
                            if cached is not None
                            else _unknown("formula mutation replay result is unavailable")
                        )
                if (
                    request.expected_revision is not None
                    and request.expected_revision != before_revision
                ):
                    self._finish_ledger(request, context, dispatched=False)
                    return _rejected(
                        otf.FormulaErrorCode.STALE_REVISION,
                        "formula target revision is stale",
                        {"revision_hash": before_revision},
                    )
                self._validate_zip(current)
                workbook = self._load_editable(current)
                temporary: str | None = None
                try:
                    worksheet = workbook[worksheet_name]
                    self._apply_formula_fill(worksheet, rectangle, request.expression)
                    workbook.calculation.fullCalcOnLoad = True
                    workbook.calculation.forceFullCalc = True
                    workbook.calculation.calcMode = "auto"
                    temporary = self._save_staged(workbook, path.parent)
                    self._assert_archive_preserved(
                        current, Path(temporary).read_bytes(), worksheet_name
                    )
                    try:
                        self._publish_staged(temporary, path)
                    except _PublicationFailure as exc:
                        replaced = exc.replaced
                        raise
                    replaced = True
                    temporary = None
                finally:
                    workbook.close()
                    if temporary is not None:
                        self._unlink_staged(temporary)
                readback_data = self._read_bytes(path)
                self._check_response_limit(readback_data, request.limits)
                readback = self._read_observation(
                    readback_data, worksheet_name, rectangle, request.limits
                )
                expected = self._expected_formula_map(rectangle, request.expression)
                observed = {cell.address: cell.expression.text for cell in readback.formulas}
                if observed != expected:
                    self._mark_unknown(request, context)
                    return _failed(
                        otf.FormulaErrorCode.READBACK_MISMATCH,
                        "formula text readback did not match the requested mutation",
                        commit=otf.FormulaCommitState.COMMITTED,
                        verification=otf.FormulaVerificationState.FAILED,
                    )
                mutation = otf.FormulaMutation(
                    target_kind="grid",
                    affected_count=rectangle.cell_count,
                    formula_observation=readback,
                    revision_before=before_revision,
                    revision_after=readback.observed_revision,
                )
                receipt = otf.FormulaReceiptDetails.for_grid_set(
                    target=request.target.grid.value,
                    selector=request.cell_range,
                    capability=otf.GRID_SET.to_reference(),
                    dialect=request.expression.dialect,
                    expression_sha256=request.expression.sha256,
                    observation_sha256=otf.formula_observation_hash(readback),
                    affected_count=rectangle.cell_count,
                    revision_before=before_revision,
                    revision_after=readback.observed_revision,
                    mutation_atomicity=otf.MutationAtomicity.ATOMIC.value,
                    revision_enforcement=otf.RevisionEnforcement.ATOMIC.value,
                    verification="formula_text_readback",
                )
                result = otf.FormulaExtensionResult(
                    value=mutation,
                    outcome=otf.FormulaOutcome.SUCCEEDED,
                    commit=otf.FormulaCommitState.COMMITTED,
                    verification=otf.FormulaVerificationState.PASSED,
                    receipts=(receipt,),
                )
                if request.idempotency_key is not None and context is not None:
                    operation_hash = _hash_bytes(str(mutation.to_wire()).encode())
                    with self._lock:
                        self._completed[operation_hash] = result
                        self._completed.move_to_end(operation_hash)
                        while len(self._completed) > self._completed_limit:
                            self._completed.popitem(last=False)
                        self._ledger.succeed(
                            connector_id=PROVIDER_EXCEL,
                            target_hash=context[0],
                            selector_hash=context[1],
                            idempotency_key=request.idempotency_key,
                            payload_hash=context[2],
                            operation_hash=operation_hash,
                        )
                return result
        except _TargetFailure as exc:
            self._finish_ledger(request, context, dispatched=replaced)
            if replaced:
                return _unknown("formula commit state could not be determined")
            return _rejected(exc.code, str(exc))
        except _LimitFailure as exc:
            self._finish_ledger(request, context, dispatched=replaced)
            if replaced:
                return _unknown("formula commit state could not be determined")
            return _rejected(otf.FormulaErrorCode.RESOURCE_LIMIT, str(exc), {"limit": exc.limit})
        except (BadZipFile, _ProtocolFailure, TypeError, ValueError, KeyError):
            self._finish_ledger(request, context, dispatched=replaced)
            if replaced:
                return _unknown("formula commit state could not be determined")
            return _failed(
                otf.FormulaErrorCode.PROTOCOL_FAILURE,
                "Excel formula mutation could not be completed",
            )
        except OSError as exc:
            if isinstance(exc, _PublicationFailure):
                replaced = exc.replaced
            self._finish_ledger(request, context, dispatched=replaced)
            if replaced:
                return _unknown("formula commit state could not be determined")
            return _failed(
                otf.FormulaErrorCode.EXECUTION_FAILED,
                "Excel formula mutation could not be published",
            )
        except Exception:
            self._finish_ledger(request, context, dispatched=replaced)
            if replaced:
                return _unknown("formula commit state could not be determined")
            return _failed(otf.FormulaErrorCode.EXECUTION_FAILED, "Excel formula mutation failed")

    def _details(self) -> otf.FormulaCapabilityDetails:
        return otf.FormulaCapabilityDetails(
            target_kind="grid",
            dialects=(otf.EXCEL_A1,),
            max_cells_per_operation=_MAX_CELLS,
            max_expression_bytes=_MAX_EXPRESSION_BYTES,
            recalculation_scopes=(),
            calculation_states=(),
            mutation_atomicity=otf.MutationAtomicity.ATOMIC,
            revision_enforcement=otf.RevisionEnforcement.ATOMIC,
            idempotency_strength=otf.IdempotencyStrength.HOST_LEDGER,
        )

    def _resolve_path(self, uri) -> tuple[Path, str | None]:
        if uri.scheme != PROVIDER_EXCEL:
            raise _TargetFailure(
                otf.FormulaErrorCode.INVALID_TARGET, "Excel Formula requires an excel:// target"
            )
        try:
            resolved = self._connector.resolve(uri, ResolveContext())
        except ConnectorError as exc:
            raise _TargetFailure(
                otf.FormulaErrorCode.INVALID_TARGET, "Excel Formula target is invalid"
            ) from exc
        path = resolved.resource.path
        if path.suffix.casefold() != ".xlsx":
            raise _TargetFailure(
                otf.FormulaErrorCode.INVALID_TARGET,
                "Excel Formula target must have an .xlsx suffix",
            )
        return path, resolved.resource.sheet

    def _assert_target_file(self, path: Path) -> None:
        if path.is_symlink():
            raise _TargetFailure(
                otf.FormulaErrorCode.INVALID_TARGET, "Excel Formula target must not be a symlink"
            )
        try:
            mode = path.stat().st_mode
        except OSError as exc:
            raise _TargetFailure(
                otf.FormulaErrorCode.INVALID_TARGET, "Excel Formula target is unavailable"
            ) from exc
        if not stat.S_ISREG(mode):
            raise _TargetFailure(
                otf.FormulaErrorCode.INVALID_TARGET, "Excel Formula target must be a regular file"
            )

    def _bound_path(self, target: otf.BoundGridFormulaTarget) -> tuple[Path, str]:
        path, uri_sheet = self._resolve_path(target.grid)
        self._assert_target_file(path)
        worksheet_id = target.worksheet.worksheet_id
        if worksheet_id is None:
            raise _TargetFailure(
                otf.FormulaErrorCode.TARGET_NOT_FOUND, "Excel worksheet binding is required"
            )
        with self._lock:
            worksheet_name = self._bindings.get((target.grid.value, worksheet_id))
        if worksheet_name is None:
            if uri_sheet is not None and uri_sheet != worksheet_id:
                raise _TargetFailure(
                    otf.FormulaErrorCode.TARGET_NOT_FOUND,
                    "Excel worksheet binding does not match the URI",
                )
            worksheet_name = worksheet_id
        if uri_sheet is not None and uri_sheet != worksheet_name:
            raise _TargetFailure(
                otf.FormulaErrorCode.TARGET_NOT_FOUND,
                "Excel worksheet binding does not match the URI",
            )
        return path, worksheet_name

    def _match_worksheet(
        self, names: list[str], reference: otf.WorksheetRef, uri_sheet: str | None
    ) -> str:
        candidates = [name for name in names if (reference.name is None or name == reference.name)]
        if reference.worksheet_id is not None:
            candidates = [name for name in candidates if name == reference.worksheet_id]
        if uri_sheet is not None:
            candidates = [name for name in candidates if name == uri_sheet]
        if len(candidates) != 1:
            raise _TargetFailure(
                otf.FormulaErrorCode.TARGET_NOT_FOUND,
                "Excel worksheet does not identify exactly one sheet",
            )
        return candidates[0]

    def _validated_range(
        self, selector: str, limits: otf.FormulaResourceLimits | None
    ) -> otf.A1Rectangle:
        rectangle = otf.A1Rectangle.parse(selector)
        requested = limits.max_cells if limits is not None else None
        if requested is not None and (
            isinstance(requested, bool) or not isinstance(requested, int) or requested <= 0
        ):
            raise _ProtocolFailure
        limit = _MAX_CELLS if requested is None else min(_MAX_CELLS, requested)
        if rectangle.cell_count > limit:
            raise _LimitFailure("formula range exceeds the configured cell limit", limit)
        return otf.A1Rectangle(
            worksheet_name=rectangle.worksheet_name,
            start_address=self._address(rectangle.start_column, rectangle.start_row),
            end_address=self._address(rectangle.end_column, rectangle.end_row),
            start_column=rectangle.start_column,
            start_row=rectangle.start_row,
            end_column=rectangle.end_column,
            end_row=rectangle.end_row,
        )

    def _expression_limit(self, limits: otf.FormulaResourceLimits | None) -> int:
        requested = limits.max_expression_bytes if limits is not None else None
        if requested is not None and (
            isinstance(requested, bool) or not isinstance(requested, int) or requested <= 0
        ):
            raise _ProtocolFailure
        return _MAX_EXPRESSION_BYTES if requested is None else min(_MAX_EXPRESSION_BYTES, requested)

    def _check_response_limit(self, data: bytes, limits: otf.FormulaResourceLimits | None) -> None:
        requested = limits.max_response_bytes if limits is not None else None
        if requested is not None and (
            isinstance(requested, bool) or not isinstance(requested, int) or requested <= 0
        ):
            raise _ProtocolFailure
        if requested is not None and len(data) > requested:
            raise _LimitFailure(
                "Excel workbook exceeds the configured response byte limit", requested
            )

    def _read_observation(
        self,
        data: bytes,
        worksheet_name: str,
        rectangle: otf.A1Rectangle,
        limits: otf.FormulaResourceLimits | None,
    ) -> otf.GridFormulaObservation:
        self._validate_zip(data)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(data), data_only=False, read_only=False, keep_links=True)
        try:
            if worksheet_name not in workbook.sheetnames:
                raise _TargetFailure(
                    otf.FormulaErrorCode.TARGET_NOT_FOUND, "Excel worksheet does not exist"
                )
            worksheet = workbook[worksheet_name]
            expression_limit = self._expression_limit(limits)
            formulas: list[otf.FormulaCell] = []
            for row in worksheet.iter_rows(
                min_row=rectangle.start_row,
                max_row=rectangle.end_row,
                min_col=rectangle.start_column,
                max_col=rectangle.end_column,
            ):
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    if not isinstance(cell.value, str):
                        raise _ProtocolFailure
                    if len(cell.value.encode("utf-8")) > expression_limit:
                        raise _LimitFailure(
                            "formula expression exceeds the configured byte limit", expression_limit
                        )
                    formulas.append(
                        otf.FormulaCell(
                            cell.coordinate, otf.FormulaExpression(cell.value, otf.EXCEL_A1)
                        )
                    )
            return otf.GridFormulaObservation(
                worksheet_id=worksheet_name,
                requested_range=self._range_text(rectangle),
                formulas=tuple(formulas),
                observed_revision=_hash_bytes(data),
            )
        finally:
            workbook.close()

    def _load_editable(self, data: bytes):
        from openpyxl import load_workbook

        return load_workbook(BytesIO(data), data_only=False, read_only=False, keep_links=True)

    def _apply_formula_fill(
        self, worksheet, rectangle: otf.A1Rectangle, expression: otf.FormulaExpression
    ) -> None:
        from openpyxl.formula.translate import Translator

        for row in range(rectangle.start_row, rectangle.end_row + 1):
            for column in range(rectangle.start_column, rectangle.end_column + 1):
                cell = worksheet.cell(row=row, column=column)
                cell.value = (
                    expression.text
                    if cell.coordinate == rectangle.start_address
                    else Translator(
                        expression.text, origin=rectangle.start_address
                    ).translate_formula(cell.coordinate)
                )

    def _expected_formula_map(
        self, rectangle: otf.A1Rectangle, expression: otf.FormulaExpression
    ) -> dict[str, str]:
        from openpyxl.formula.translate import Translator

        return {
            self._address(column, row): (
                expression.text
                if self._address(column, row) == rectangle.start_address
                else Translator(expression.text, origin=rectangle.start_address).translate_formula(
                    self._address(column, row)
                )
            )
            for row in range(rectangle.start_row, rectangle.end_row + 1)
            for column in range(rectangle.start_column, rectangle.end_column + 1)
        }

    def _read_bytes(self, path: Path) -> bytes:
        self._assert_target_file(path)
        flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
        descriptor = os.open(path, flags)
        try:
            current = os.fstat(descriptor)
            if not stat.S_ISREG(current.st_mode):
                raise OSError("Excel Formula target is not a regular file")
            with os.fdopen(descriptor, "rb", closefd=False) as stream:
                return stream.read()
        finally:
            os.close(descriptor)

    def _validate_zip(self, data: bytes) -> None:
        try:
            with ZipFile(BytesIO(data)) as archive:
                names = archive.namelist()
                if len(names) != len(set(names)) or any(
                    name.startswith(("/", "\\")) or ".." in Path(name).parts for name in names
                ):
                    raise _ProtocolFailure
                if archive.testzip() is not None:
                    raise _ProtocolFailure
        except BadZipFile:
            raise
        except OSError as exc:
            raise _ProtocolFailure from exc

    def _assert_archive_preserved(
        self, original: bytes, staged: bytes, worksheet_name: str
    ) -> None:
        try:
            with ZipFile(BytesIO(original)) as before, ZipFile(BytesIO(staged)) as after:
                before_names = set(before.namelist())
                after_names = set(after.namelist())
                if before_names != after_names:
                    raise _ProtocolFailure
                worksheet_part = self._worksheet_archive_name(before, worksheet_name)
                allowed = {"xl/workbook.xml", worksheet_part}
                for name in before_names - allowed:
                    if before.read(name) != after.read(name):
                        raise _ProtocolFailure
        except (BadZipFile, KeyError, OSError) as exc:
            raise _ProtocolFailure from exc

    def _worksheet_archive_name(self, archive: ZipFile, worksheet_name: str) -> str:
        workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_namespace = "{http://schemas.openxmlformats.org/package/2006/relationships}"
        rel_targets = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in relationships
            if relationship.tag == f"{rel_namespace}Relationship"
        }
        sheet_namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        document_namespace = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
        for sheet in workbook_xml.findall(f"{sheet_namespace}sheets/{sheet_namespace}sheet"):
            if sheet.attrib.get("name") == worksheet_name:
                target = rel_targets.get(sheet.attrib.get(f"{document_namespace}id", ""))
                if target is not None:
                    target = target.lstrip("/")
                    if not target.startswith("xl/"):
                        target = posixpath.join("xl", target)
                    return posixpath.normpath(target)
        raise _ProtocolFailure

    def _save_staged(self, workbook, directory: Path) -> str:
        descriptor, temporary = tempfile.mkstemp(
            prefix=".otc-excel-", suffix=".xlsx", dir=directory
        )
        os.close(descriptor)
        try:
            os.chmod(temporary, 0o600)
            workbook.save(temporary)
            with open(temporary, "rb") as stream:
                stream.flush()
                os.fsync(stream.fileno())
            return temporary
        except Exception:
            self._unlink_staged(temporary)
            raise

    def _publish_staged(self, temporary: str, path: Path) -> None:
        try:
            os.replace(temporary, path)
        except OSError as exc:
            raise _PublicationFailure(str(exc), replaced=False) from exc
        try:
            directory_descriptor = os.open(path.parent, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0))
        except OSError as exc:
            raise _PublicationFailure(str(exc), replaced=True) from exc
        try:
            try:
                os.fsync(directory_descriptor)
            except OSError as exc:
                raise _PublicationFailure(str(exc), replaced=True) from exc
        finally:
            os.close(directory_descriptor)

    def _range_text(self, rectangle: otf.A1Rectangle) -> str:
        if rectangle.start_address == rectangle.end_address:
            return rectangle.start_address
        return f"{rectangle.start_address}:{rectangle.end_address}"

    @staticmethod
    def _address(column: int, row: int) -> str:
        letters = ""
        value = column
        while value:
            value, remainder = divmod(value - 1, 26)
            letters = chr(ord("A") + remainder) + letters
        return f"{letters}{row}"

    def _unlink_staged(self, temporary: str) -> None:
        with suppress(FileNotFoundError):
            os.unlink(temporary)

    @contextmanager
    def _target_lock(self, path: Path, lock_path: Path):
        import fcntl

        self._assert_target_file(path)
        descriptor = os.open(
            lock_path,
            os.O_RDWR | os.O_CREAT | getattr(os, "O_NOFOLLOW", 0),
            0o600,
        )
        try:
            if not stat.S_ISREG(os.fstat(descriptor).st_mode):
                raise OSError("Excel Formula lock is not a regular file")
            fcntl.flock(descriptor, fcntl.LOCK_EX)
            self._assert_target_file(path)
            yield None
        finally:
            fcntl.flock(descriptor, fcntl.LOCK_UN)
            os.close(descriptor)

    def _finish_ledger(
        self, request, context: tuple[str, str, str] | None, *, dispatched: bool
    ) -> None:
        if request.idempotency_key is None or context is None:
            return
        try:
            with self._lock:
                if dispatched:
                    self._ledger.mark_unknown(
                        connector_id=PROVIDER_EXCEL,
                        target_hash=context[0],
                        selector_hash=context[1],
                        idempotency_key=request.idempotency_key,
                        payload_hash=context[2],
                    )
                else:
                    self._ledger.fail_known(
                        connector_id=PROVIDER_EXCEL,
                        target_hash=context[0],
                        selector_hash=context[1],
                        idempotency_key=request.idempotency_key,
                        payload_hash=context[2],
                    )
        except (KeyError, ValueError):
            pass

    def _mark_unknown(self, request, context: tuple[str, str, str] | None) -> None:
        if request.idempotency_key is None or context is None:
            return
        try:
            with self._lock:
                self._ledger.mark_unknown(
                    connector_id=PROVIDER_EXCEL,
                    target_hash=context[0],
                    selector_hash=context[1],
                    idempotency_key=request.idempotency_key,
                    payload_hash=context[2],
                )
        except (KeyError, ValueError):
            pass


class _TargetFailure(ValueError):
    def __init__(self, code: otf.FormulaErrorCode, message: str) -> None:
        super().__init__(message)
        self.code = code


__all__ = ["ExcelFormulaExtension"]

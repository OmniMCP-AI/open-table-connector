from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from open_table_connector.cli.model import CliOptions, FormatName, parse_endpoint
from open_table_connector.cli.pipeline import convert_endpoint, import_endpoint
from open_table_connector.cli.registry import build_default_registry
from open_table_connector.contract import ConnectorError, ConnectorErrorCode, TableMode


def test_cli_lists_concrete_local_connector_types() -> None:
    registry = build_default_registry()

    identities = {adapter.identity.connector_id for adapter in registry.list()}

    assert {"local_files", "csv", "excel", "md"} <= identities


@pytest.mark.parametrize(
    ("raw", "connector_id"),
    (
        ("csv:///tmp/orders.csv", "csv"),
        ("excel:///tmp/orders.xlsx", "excel"),
        ("md:///tmp/orders.md", "md"),
    ),
)
def test_registry_routes_explicit_local_scheme(raw: str, connector_id: str) -> None:
    adapter = build_default_registry().connector_for(parse_endpoint(raw))

    assert adapter.identity.connector_id == connector_id


def test_registry_routes_bare_path_to_local_files_facade(tmp_path: Path) -> None:
    endpoint = parse_endpoint(str(tmp_path / "orders.csv"))

    adapter = build_default_registry().connector_for(endpoint)

    assert adapter.identity.connector_id == "local_files"


def test_local_adapter_auto_probes_extensionless_csv_and_preserves_facade_receipt(
    tmp_path: Path,
) -> None:
    source = tmp_path / "orders"
    source.write_text("id,note\n1,ok\n", encoding="utf8")
    endpoint = parse_endpoint(str(source))

    result = build_default_registry().connector_for(endpoint).read(endpoint, CliOptions())

    assert result.table.to_pylist() == [{"id": "1", "note": "ok"}]
    assert result.receipt.connector.connector_id == "local_files"
    assert result.receipt.mode is TableMode.SHEET


def test_local_adapter_auto_probes_content_despite_misleading_suffix(tmp_path: Path) -> None:
    source = tmp_path / "orders.xlsx"
    source.write_text("id,note\n1,ok\n", encoding="utf8")
    endpoint = parse_endpoint(source.as_uri())

    result = build_default_registry().connector_for(endpoint).read(endpoint, CliOptions())

    assert result.table.to_pylist() == [{"id": "1", "note": "ok"}]


def test_local_adapter_auto_honors_requested_excel_sheet(tmp_path: Path) -> None:
    source = tmp_path / "orders"
    workbook = Workbook()
    orders = workbook.active
    orders.title = "orders"
    orders.append(["id"])
    orders.append(["1"])
    refunds = workbook.create_sheet("refunds")
    refunds.append(["refund_id"])
    refunds.append(["r1"])
    workbook.save(source)
    endpoint = parse_endpoint(str(source))

    result = build_default_registry().connector_for(endpoint).read(
        endpoint,
        CliOptions(sheet="refunds"),
    )

    assert result.table.to_pylist() == [{"refund_id": "r1"}]
    assert result.receipt.connector.connector_id == "local_files"
    assert result.receipt.coordinate_convention.sheet == "refunds"


def test_local_adapter_auto_inspection_reports_native_sheet_facts(tmp_path: Path) -> None:
    source = tmp_path / "orders"
    source.write_text("| id |\n| --- |\n| 1 |\n", encoding="utf8")
    endpoint = parse_endpoint(str(source))

    inspection = build_default_registry().connector_for(endpoint).inspect(
        endpoint,
        CliOptions(),
    )

    assert inspection.mode is TableMode.SHEET
    assert inspection.coordinate_convention.sheet == "data"
    assert inspection.facts["worksheets"] == ["data"]


@pytest.mark.parametrize(
    ("format_name", "payload"),
    (
        (FormatName.JSON, '[{"id":"1"}]'),
        (FormatName.JSONL, '{"id":"1"}\n'),
    ),
)
def test_local_adapter_retains_explicit_json_reading(
    tmp_path: Path,
    format_name: FormatName,
    payload: str,
) -> None:
    source = tmp_path / "orders.data"
    source.write_text(payload, encoding="utf8")
    endpoint = parse_endpoint(str(source))

    result = build_default_registry().connector_for(endpoint).read(
        endpoint,
        CliOptions(from_format=format_name),
    )

    assert result.table.to_pylist() == [{"id": "1"}]
    assert result.receipt.connector.connector_id == "local_files"


def test_excel_adapter_inspection_delegates_native_sheet_facts(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    workbook = Workbook()
    orders = workbook.active
    orders.title = "orders"
    orders.append(["id"])
    orders.append(["1"])
    refunds = workbook.create_sheet("refunds")
    refunds.append(["refund_id"])
    refunds.append(["r1"])
    workbook.save(source)
    endpoint = parse_endpoint(f"excel://{source}")

    inspection = build_default_registry().connector_for(endpoint).inspect(
        endpoint,
        CliOptions(sheet="refunds"),
    )

    assert inspection.mode is TableMode.SHEET
    assert inspection.columns == ("refund_id",)
    assert inspection.coordinate_convention.sheet == "refunds"
    assert inspection.coordinate_convention.header_rows == 1
    assert inspection.coordinate_convention.first_data_row == 2
    assert inspection.facts == {
        "worksheets": ["orders", "refunds"],
        "formula_text_captured": False,
        "formula_calculated": False,
    }


def test_cli_converts_csv_to_explicit_markdown_destination(tmp_path: Path) -> None:
    source = tmp_path / "orders.csv"
    destination = tmp_path / "orders.md"
    source.write_text("id\n1\n", encoding="utf8")

    summary = convert_endpoint(
        parse_endpoint(str(source)),
        parse_endpoint(f"md://{destination}"),
        build_default_registry(),
        CliOptions(),
    )

    assert summary.rows_written == 1
    assert destination.read_text(encoding="utf8").splitlines()[0].startswith("| id")


def test_cli_converts_csv_to_explicit_excel_destination(tmp_path: Path) -> None:
    source = tmp_path / "orders.csv"
    destination = tmp_path / "orders.xlsx"
    source.write_text("id\n1\n", encoding="utf8")

    summary = convert_endpoint(
        parse_endpoint(str(source)),
        parse_endpoint(f"excel://{destination}"),
        build_default_registry(),
        CliOptions(),
    )

    assert summary.rows_written == 1
    workbook = load_workbook(destination, read_only=True, data_only=True)
    try:
        assert list(workbook.active.values) == [("id",), ("1",)]
    finally:
        workbook.close()


def test_explicit_local_destination_scheme_takes_precedence_over_output_format(tmp_path: Path) -> None:
    source = tmp_path / "orders.csv"
    destination = tmp_path / "orders.md"
    source.write_text("id\n1\n", encoding="utf8")

    convert_endpoint(
        parse_endpoint(str(source)),
        parse_endpoint(f"md://{destination}"),
        build_default_registry(),
        CliOptions(output_format=FormatName.JSON),
    )

    assert destination.read_text(encoding="utf8").splitlines()[0].startswith("| id")


def test_import_rejects_explicit_local_destination_before_read(tmp_path: Path) -> None:
    source = tmp_path / "orders.csv"
    source.write_text("id\n1\n", encoding="utf8")

    with pytest.raises(ConnectorError) as error:
        import_endpoint(
            parse_endpoint(str(source)),
            parse_endpoint(f"csv://{tmp_path / 'copy.csv'}"),
            build_default_registry(),
            CliOptions(),
        )

    assert error.value.code is ConnectorErrorCode.UNSUPPORTED_CAPABILITY
    assert error.value.safe_details["scheme"] == "csv"
